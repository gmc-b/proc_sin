import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment

def read_acq(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    
    # Extract metadata
    for line in lines:
        if "msec/sample" in line:
            sample_rate = float(line.split()[0]) / 1000  # msec/sample
            break

    # Find the line where data starts (after headers)
    for i, line in enumerate(lines):
        if line.startswith("sec"):
            header_line = i
            break
    
    # Extract column names
    headers = lines[header_line].strip().split('\t')
    # Load data
    data = pd.read_csv(file_path, delimiter='\t', skiprows=header_line + 2, names=headers, na_values=[''], index_col=False)
    return data, sample_rate

def find_oscilation(signal, time, ch, peak_prominence, peak_distance,save_dir=None):
    # Find peaks
    peak_indices, _ = find_peaks(signal, prominence=peak_prominence, distance=peak_distance)
    
    # Find valleys
    valley_indices, _ = find_peaks(-signal, prominence=peak_prominence, distance=peak_distance)
    
    # Pair peaks and valleys based on the smallest time difference
    oscilation_pairs = []
    for peak_idx in peak_indices:
        time_differences = np.abs(time[valley_indices] - time[peak_idx])
        closest_valley_idx = valley_indices[np.argmin(time_differences)]
        oscilation_pairs.append((peak_idx, closest_valley_idx))

    # Extract the first 5 pairs
    first_5_pairs = oscilation_pairs[:5]
    
    # Calculate max height, min height, and amplitude
    results = []
    for peak_idx, valley_idx in first_5_pairs:
        max_height = signal[peak_idx]
        min_height = signal[valley_idx]
        amplitude = max_height - min_height
        results.append((max_height, min_height, amplitude))

    if save_dir:
        plt.figure(figsize=(10, 6))
        plt.plot(time, signal, label='Signal')
        plt.plot(time[peak_indices], signal[peak_indices], 'ro', label='Detected Peaks')
        plt.plot(time[valley_indices], signal[valley_indices], 'go', label='Detected Valleys')
        
        # Highlight paired peaks and valleys
        for pair in oscilation_pairs:
            plt.plot([time[pair[0]], time[pair[1]]], [signal[pair[0]], signal[pair[1]]], 'k--', alpha=0.5)  # Connect pairs with a dashed line

        plt.xlabel('Time (s)')
        plt.ylabel('Amplitude')
        plt.title(f'Detected Peaks and Valleys for {ch}')
        plt.legend()
        plt.savefig(os.path.join(save_dir, f"{ch}.png"))
        plt.close()

    return results

def read_parameters(file_path):
    try:
        with open(file_path, 'r') as file:
            parameters = json.load(file)
        return parameters
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return {}
    except json.JSONDecodeError:
        print(f"Error: The file '{file_path}' is not a valid JSON file.")
        return {}

def validate_channels(json_channels, file_channels):
    valid_channels = []
    for jc in json_channels:
        if jc in file_channels:
            valid_channels.append(jc)
        else:
            print(f"Canal [{jc}] não existe no arquivo .acq e será ignorado")
    return valid_channels

def process_file(file_path, param, output_data, output_folder):
    data, sample_rate = read_acq(file_path)
    valid_channels = validate_channels(param.keys(), data.columns.tolist())
    time = data["sec"]

    # Create a subfolder for the current file
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    file_output_dir = os.path.join(output_folder, file_name)
    os.makedirs(file_output_dir, exist_ok=True)

    for ch in valid_channels:
        signal = data[ch]
        channel = param[ch]
        peak_prominence = channel["peak_prominence"]
        peak_min_distance = channel["peak_min_distance"]

        # Find oscillations and extract the first 5 pairs
        results = find_oscilation(signal, time, ch, peak_prominence, peak_min_distance, save_dir=file_output_dir)
        
        # Store results in the output DataFrame
        for i, (max_height, min_height, amplitude) in enumerate(results):
            output_data.append({
                "File": file_name,
                "Channel": ch,
                "Pair": i + 1,
                "Max": max_height,
                "Min": min_height,
                "Amplitude": amplitude
            })

def save_to_spreadsheet(output_data, output_csv):
    # Create a DataFrame from the output data
    output_df = pd.DataFrame(output_data)
    output_df = output_df[["File", "Channel", "Pair", "Max", "Min", "Amplitude"]]

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Add headers
    headers = ["File", "Channel", "Pair", "Max", "Min", "Amplitude"]
    ws.append(headers)

    # Add data and merge cells for the File and Channel columns
    current_file = None
    current_channel = None
    file_start_row = 2
    channel_start_row = 2

    for row in output_df.itertuples(index=False):
        if row.File != current_file:
            # Merge cells for the previous file
            if current_file is not None:
                ws.merge_cells(start_row=file_start_row, start_column=1, end_row=ws.max_row, end_column=1)
                ws.cell(row=file_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            current_file = row.File
            file_start_row = ws.max_row + 1

        if row.Channel != current_channel:
            # Merge cells for the previous channel
            if current_channel is not None:
                ws.merge_cells(start_row=channel_start_row, start_column=2, end_row=ws.max_row, end_column=2)
                ws.cell(row=channel_start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
            current_channel = row.Channel
            channel_start_row = ws.max_row + 1

        # Append the row data
        ws.append([row.File, row.Channel, row.Pair, row.Max, row.Min, row.Amplitude])

    # Merge cells for the last file and channel
    if current_file is not None:
        ws.merge_cells(start_row=file_start_row, start_column=1, end_row=ws.max_row, end_column=1)
        ws.cell(row=file_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    if current_channel is not None:
        ws.merge_cells(start_row=channel_start_row, start_column=2, end_row=ws.max_row, end_column=2)
        ws.cell(row=channel_start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook
    wb.save(output_csv)

def main():
    data_folder = "Data"
    param_path = "parametros.json"
    output_folder = "Output"
    output_csv = os.path.join(output_folder, "resultados.xlsx")

    # Create the Output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)

    # Read parameters from JSON
    param = read_parameters(param_path)
    if not param:
        return

    # Initialize output data storage
    output_data = []

    # Process all .acq files in the Data folder
    for file_name in os.listdir(data_folder):
        if file_name.endswith(".acq"):
            file_path = os.path.join(data_folder, file_name)
            print(f"Processing file: {file_name}")
            process_file(file_path, param, output_data, output_folder)

    # Save results to a spreadsheet
    if output_data:
        save_to_spreadsheet(output_data, output_csv)
        print(f"Results saved to {output_csv}")
    else:
        print("No data to save.")

if __name__ == "__main__":
    main()